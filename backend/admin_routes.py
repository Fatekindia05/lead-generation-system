from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import JSONResponse, StreamingResponse
from typing import Optional
import io
import csv
import base64
from datetime import datetime
import os
import tempfile
import shutil

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from PIL import Image

from models import LeadUpdate, LeadsFilter
from lead_service import (
    get_all_leads,
    get_lead_by_id,
    update_lead_status,
    get_leads_statistics,
    delete_lead,
    export_leads
)

router = APIRouter(prefix="/api/admin", tags=["Admin"])

@router.get("/leads")
async def get_leads(
    status: Optional[str] = Query(None, description="Filter by status"),
    requirement_type: Optional[str] = Query(None, description="Filter by requirement type"),
    customer_type: Optional[str] = Query(None, description="Filter by customer type"),
    search: Optional[str] = Query(None, description="Search in name, email, company"),
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0)
):
    """Get all leads with filters"""
    leads = get_all_leads(
        status=status,
        requirement_type=requirement_type,
        customer_type=customer_type,
        search=search,
        limit=limit,
        offset=offset
    )
    return {"leads": leads, "total": len(leads)}

@router.get("/leads/{lead_id}")
async def get_lead(lead_id: int):
    """Get lead by ID"""
    lead = get_lead_by_id(lead_id)
    if not lead:
        raise HTTPException(status_code=404, detail="Lead not found")
    return lead

@router.patch("/leads/{lead_id}/status")
async def update_lead(lead_id: int, update: LeadUpdate):
    """Update lead status"""
    success = update_lead_status(lead_id, update.status)
    if not success:
        raise HTTPException(status_code=404, detail="Lead not found")
    return {"status": "success", "message": f"Lead {lead_id} updated to {update.status}"}

@router.delete("/leads/{lead_id}")
async def remove_lead(lead_id: int):
    """Delete lead by ID"""
    success = delete_lead(lead_id)
    if not success:
        raise HTTPException(status_code=404, detail="Lead not found")
    return {"status": "success", "message": f"Lead {lead_id} deleted"}

@router.get("/stats")
async def get_stats():
    """Get lead statistics"""
    return get_leads_statistics()

@router.get("/export/csv")
async def export_csv():
    """Export leads as CSV"""
    leads = export_leads()
    
    if not leads:
        raise HTTPException(status_code=404, detail="No leads to export")
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    headers = ["ID", "Name", "Company", "Email", "Phone", "Requirement Type", 
               "Customer Type", "Other Customer Type", "Status", "Created At", "Message"]
    writer.writerow(headers)
    
    for lead in leads:
        writer.writerow([
            lead.get('id'),
            lead.get('name'),
            lead.get('company'),
            lead.get('email'),
            lead.get('phone'),
            lead.get('requirement_type'),
            lead.get('customer_type'),
            lead.get('other_customer_type', ''),
            lead.get('status'),
            lead.get('created_at'),
            lead.get('message', '')
        ])
    
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename=leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        }
    )

@router.get("/export/excel")
async def export_excel():
    """
    Export leads as Excel file with embedded images
    Cell size exactly matches image size
    """
    leads = export_leads()
    
    if not leads:
        raise HTTPException(status_code=404, detail="No leads to export")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers with Image column
    headers = ["ID", "Image", "Name", "Company", "Email", "Phone", 
               "Requirement Type", "Customer Type", "Other Customer Type", 
               "Status", "Created At", "Message"]
    
    # Column widths - set initial widths
    col_widths = [6, 40, 20, 25, 30, 15, 20, 18, 20, 12, 22, 50]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col)].width = col_widths[col-1]
    
    ws.row_dimensions[1].height = 30
    
    # Track the maximum image width for column adjustment
    max_image_width = 0
    
    # Store all image details to calculate final column width
    image_details = []
    
    # First pass: Process all images to get dimensions
    for row_idx, lead in enumerate(leads, 2):
        image_url = lead.get('image_url')
        image_detail = {
            'row_idx': row_idx,
            'lead': lead,
            'image_path': None,
            'width': 0,
            'height': 0,
            'has_image': False
        }
        
        if image_url:
            try:
                filename = image_url.split('/')[-1]
                image_path = os.path.join("data/images", filename)
                
                if os.path.exists(image_path):
                    with Image.open(image_path) as img:
                        # Convert to RGB if necessary
                        if img.mode in ('RGBA', 'LA', 'P'):
                            img = img.convert('RGB')
                        
                        # Target size: maintain aspect ratio, max 350x280 pixels
                        max_width = 350
                        max_height = 280
                        
                        # Calculate aspect ratio
                        width_ratio = max_width / img.width
                        height_ratio = max_height / img.height
                        ratio = min(width_ratio, height_ratio, 1.0)  # Don't upscale
                        
                        new_width = int(img.width * ratio)
                        new_height = int(img.height * ratio)
                        
                        image_detail['has_image'] = True
                        image_detail['width'] = new_width
                        image_detail['height'] = new_height
                        image_detail['image_path'] = image_path
                        
                        # Track max width for column adjustment
                        if new_width > max_image_width:
                            max_image_width = new_width
            except Exception as e:
                print(f"Error processing image for lead {lead.get('id')}: {e}")
        
        image_details.append(image_detail)
    
    # Calculate column width based on max image width
    # Excel column width: 1 unit = 7 pixels approximately
    if max_image_width > 0:
        # Add 5 pixels padding for cell borders
        image_column_width = (max_image_width + 10) / 7
        # Ensure minimum width
        image_column_width = max(image_column_width, 25)
        # Set the image column (column B) width
        ws.column_dimensions['B'].width = image_column_width
    
    # Create temp directory for resized images
    temp_dir = tempfile.mkdtemp()
    temp_files = []
    
    try:
        # Second pass: Add images and set row heights
        for detail in image_details:
            row_idx = detail['row_idx']
            lead = detail['lead']
            
            # ID
            ws.cell(row=row_idx, column=1, value=lead.get('id'))
            ws.cell(row=row_idx, column=1).alignment = center_alignment
            
            # Image
            if detail['has_image']:
                try:
                    img = Image.open(detail['image_path'])
                    if img.mode in ('RGBA', 'LA', 'P'):
                        img = img.convert('RGB')
                    
                    # Resize to calculated dimensions
                    img_resized = img.resize(
                        (detail['width'], detail['height']), 
                        Image.Resampling.LANCZOS
                    )
                    
                    # Save to temp file with high quality
                    temp_path = os.path.join(temp_dir, f"img_{row_idx}.jpg")
                    img_resized.save(temp_path, 'JPEG', quality=95, optimize=True)
                    temp_files.append(temp_path)
                    
                    # Add to Excel
                    xl_img = XLImage(temp_path)
                    # Set exact pixel dimensions in Excel
                    xl_img.width = detail['width']
                    xl_img.height = detail['height']
                    
                    cell_ref = f'B{row_idx}'
                    ws.add_image(xl_img, cell_ref)
                    
                    # Set row height to match image + padding (pixels to points)
                    # 1 point = 1.333 pixels (Excel uses points for row height)
                    row_height = (detail['height'] + 10) / 1.333
                    ws.row_dimensions[row_idx].height = row_height
                    
                except Exception as e:
                    print(f"Error adding image for lead {lead.get('id')}: {e}")
                    ws.cell(row=row_idx, column=2, value="Image Error")
                    ws.cell(row=row_idx, column=2).alignment = center_alignment
                    ws.row_dimensions[row_idx].height = 20
            else:
                ws.cell(row=row_idx, column=2, value="No Image")
                ws.cell(row=row_idx, column=2).alignment = center_alignment
                ws.row_dimensions[row_idx].height = 20
            
            # Other fields
            ws.cell(row=row_idx, column=3, value=lead.get('name'))
            ws.cell(row=row_idx, column=4, value=lead.get('company'))
            ws.cell(row=row_idx, column=5, value=lead.get('email'))
            ws.cell(row=row_idx, column=6, value=lead.get('phone', ''))
            ws.cell(row=row_idx, column=7, value=lead.get('requirement_type'))
            ws.cell(row=row_idx, column=8, value=lead.get('customer_type'))
            ws.cell(row=row_idx, column=9, value=lead.get('other_customer_type', ''))
            ws.cell(row=row_idx, column=10, value=lead.get('status'))
            ws.cell(row=row_idx, column=11, value=lead.get('created_at'))
            ws.cell(row=row_idx, column=12, value=lead.get('message', ''))
            ws.cell(row=row_idx, column=12).alignment = wrap_alignment
        
        # Auto-adjust other column widths
        for col in [1, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]:
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(2, len(leads) + 2):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    try:
                        if len(str(cell_value)) > max_length:
                            max_length = len(str(cell_value))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            if adjusted_width > ws.column_dimensions[column_letter].width:
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=leads_with_images_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )
    
    finally:
        # Clean up temp files
        shutil.rmtree(temp_dir, ignore_errors=True)